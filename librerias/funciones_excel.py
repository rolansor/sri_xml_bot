from openpyxl import Workbook
from datetime import datetime
from librerias.diccionarios import nombres_comprobantes_excel


def adjust_column_width(ws):
    """
    Ajusta el ancho de las columnas de una hoja de cálculo de Excel basado en la longitud del contenido de cada celda.
    Parámetros:
        ws (Worksheet): La hoja de cálculo de Excel a ajustar.
    No retorna nada.
    """
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells)
        index = column_cells[0].column_letter
        ws.column_dimensions[index].width = max_length + 2  # Ajuste adicional


# Esta función solo se utiliza una vez hay que buscar como eliminarla.
def formatear_fecha(fecha_str):
    try:
        return datetime.strptime(fecha_str, "%d/%m/%Y").date()
    except ValueError:
        return fecha_str


def aplanar_diccionario_nc(d, items=None, omitir_campos=None):
    if omitir_campos is None:
        omitir_campos = ['detalles', 'infoAdicional']
    if items is None:
        # Inicializar todas las columnas de IVA en 0
        items = {'Base 0%': 0, 'Base 5%': 0, 'Base 12%': 0, 'Base 14%': 0, 'Base 15%': 0, 'Base No Objeto': 0, 'Base Exento': 0,
                 'IVA 0%': 0, 'IVA 5%': 0, 'IVA 12%': 0, 'IVA 14%': 0, 'IVA 15%': 0, 'IVA No Objeto': 0, 'Exento': 0}
    for k, v in d.items():
        if k in omitir_campos:
            continue

        if k == 'totalConImpuestos' and isinstance(v, list):
            for impuesto_info in v:
                codigo = impuesto_info.get('codigo', '')
                codigoPorcentaje = impuesto_info.get('codigoPorcentaje', '')
                if codigo == '2':  # IVA
                    if codigoPorcentaje == '0':
                        items['Base 0%'] += float(impuesto_info.get('baseImponible', 0))
                        items['IVA 0%'] += float(impuesto_info.get('valor', 0))
                    elif codigoPorcentaje == '2':
                        items['Base 12%'] += float(impuesto_info.get('baseImponible', 0))
                        items['IVA 12%'] += float(impuesto_info.get('valor', 0))
                    elif codigoPorcentaje == '3':
                        items['Base 14%'] += float(impuesto_info.get('baseImponible', 0))
                        items['IVA 14%'] += float(impuesto_info.get('valor', 0))
                    elif codigoPorcentaje == '4':
                        items['Base 15%'] += float(impuesto_info.get('baseImponible', 0))
                        items['IVA 15%'] += float(impuesto_info.get('valor', 0))
                    elif codigoPorcentaje == '5':
                        items['Base 5%'] += float(impuesto_info.get('baseImponible', 0))
                        items['IVA 5%'] += float(impuesto_info.get('valor', 0))
                    elif codigoPorcentaje == '6':
                        items['Base No Objeto'] += float(impuesto_info.get('baseImponible', 0))
                        items['IVA No Objeto'] += float(impuesto_info.get('valor', 0))
                    elif codigoPorcentaje == '7':
                        items['Base Exento'] += float(impuesto_info.get('baseImponible', 0))
                        items['Exento'] += float(impuesto_info.get('valor', 0))
            continue

        if k == 'identificacionComprador':
            items['ruc_receptor'] = v
        if k == 'ruc':
            items['ruc_emisor'] = v

        if isinstance(v, dict):
            items.update(aplanar_diccionario_nc(v, items, omitir_campos))
        else:
            items[k] = v
    return items


def aplanar_diccionario_ret(d, items=None, omitir_campos=None):
    if omitir_campos is None:
        omitir_campos = ['maquinaFiscal', 'infoAdicional']
    if items is None:
        # Inicializar todas las columnas de retenciones en 0
        items = {'Renta': 0, 'IVA': 0, 'ISD': 0}
    for k, v in d.items():
        if k in omitir_campos:
            continue

        if k == 'impuestos' and isinstance(v, list):
            for impuesto_info in v:
                codigo = impuesto_info.get('codigo', '')
                valorRetenido = float(impuesto_info.get('valorRetenido', 0))
                if codigo == '1':
                    items['Renta'] += valorRetenido
                elif codigo == '2':
                    items['IVA'] += valorRetenido
                elif codigo == '6':
                    items['ISD'] += valorRetenido
            continue

        if k == 'identificacionSujetoRetenido':
            items['ruc_receptor'] = v
        if k == 'ruc':
            items['ruc_emisor'] = v

        if isinstance(v, dict):
            items.update(aplanar_diccionario_ret(v, items, omitir_campos))
        else:
            items[k] = v
    return items


def aplanar_diccionario_fac(d, items=None, omitir_campos=None):
    if items is None:
        items = {
            'Base 0%': 0, 'Base 5%': 0, 'Base 12%': 0, 'Base 14%': 0, 'Base 15%': 0, 'Base No Objeto': 0, 'Base Exento': 0,
            'IVA 0%': 0, 'IVA 5%': 0, 'IVA 12%': 0, 'IVA 14%': 0, 'IVA 15%': 0, 'IVA No Objeto': 0, 'IVA Exento': 0
        }

    if omitir_campos is None:
        omitir_campos = ['infoAdicional', 'detalles']

    for k, v in d.items():
        if k in omitir_campos:
            continue

        if k == 'totalConImpuestos' and isinstance(v, list):
            for impuesto_info in v:
                codigo = impuesto_info.get('codigo', '')
                codigoPorcentaje = impuesto_info.get('codigoPorcentaje', '')
                baseImponible = float(impuesto_info.get('baseImponible', 0))
                valor = float(impuesto_info.get('valor', 0))
                if codigo == '2':  # IVA
                    if codigoPorcentaje == '0':
                        items['Base 0%'] += float(impuesto_info.get('baseImponible', 0))
                        items['IVA 0%'] += float(impuesto_info.get('valor', 0))
                    elif codigoPorcentaje == '2':
                        items['Base 12%'] += float(impuesto_info.get('baseImponible', 0))
                        items['IVA 12%'] += float(impuesto_info.get('valor', 0))
                    elif codigoPorcentaje == '3':
                        items['Base 14%'] += float(impuesto_info.get('baseImponible', 0))
                        items['IVA 14%'] += float(impuesto_info.get('valor', 0))
                    elif codigoPorcentaje == '4':
                        items['Base 15%'] += float(impuesto_info.get('baseImponible', 0))
                        items['IVA 15%'] += float(impuesto_info.get('valor', 0))
                    elif codigoPorcentaje == '5':
                        items['Base 5%'] += float(impuesto_info.get('baseImponible', 0))
                        items['IVA 5%'] += float(impuesto_info.get('valor', 0))
                    elif codigoPorcentaje == '6':
                        items['Base No Objeto'] += float(impuesto_info.get('baseImponible', 0))
                        items['IVA No Objeto'] += float(impuesto_info.get('valor', 0))
                    elif codigoPorcentaje == '7':
                        items['Base Exento'] += float(impuesto_info.get('baseImponible', 0))
                        items['IVA Exento'] += float(impuesto_info.get('valor', 0))
            continue

        if k == 'identificacionComprador':
            items['ruc_receptor'] = v
        if k == 'ruc':
            items['ruc_emisor'] = v

        if isinstance(v, dict):
            items.update(aplanar_diccionario_fac(v, items, omitir_campos))
        else:
            items[k] = v
    return items


def guardar_documentos_emitidos(diccionario_documento, file_path):
    # Inicialización del diccionario para almacenar datos clasificados por tipo de comprobante
    sheets_data = {
        "Facturas": [],
        "Comprobantes de Retención": [],
        "Notas de Crédito": []
    }

    wb = Workbook()
    sheet_index = 0  # Inicia un contador para las hojas de trabajo

    # Orden de las columnas para Notas de Crédito y Comprobantes de Retención
    column_order_nc = [
        "codDoc", "ruc_emisor", "razonSocial", "contribuyenteEspecial",
        "obligadoContabilidad", "estab", "ptoEmi", "secuencial",
        "claveAcceso", "fechaEmision", "codDocModificado", "numDocModificado",
        "fechaEmisionDocSustento", "totalSinImpuestos",
        "Base 0%", "Base 5%", "Base 12%", "Base 15%", "Base No Objeto", "Base Exento", "IVA 5%", "IVA 12%", "IVA 15%", "valorModificacion",
        "ruc_receptor", "razonSocialComprador"
    ]

    column_order_ret = [
        "codDoc", "ruc_emisor", "razonSocial",
        "agenteRetencion", "estab", "ptoEmi", "secuencial",
        "claveAcceso", "fechaEmision", "periodoFiscal", "Renta", "IVA", "ISD",  "ruc_receptor", "razonSocialSujetoRetenido"
    ]

    column_order_fac = [
        "codDoc", "ruc_emisor", "razonSocial", "contribuyenteEspecial",
        "estab", "ptoEmi", "secuencial",
        "claveAcceso", "fechaEmision", "totalSinImpuestos", "totalDescuento",
        "Base 0%", "Base 5%", "Base 12%", "Base 15%", "Base No Objeto", "Base Exento", "IVA 5%", "IVA 12%", "IVA 15%", "importeTotal",
        "ruc_receptor", "razonSocialComprador"
    ]

    # Clasificación de documentos por tipo de comprobante
    for documento in diccionario_documento:
        codDoc = documento.get("infoTributaria", {}).get("codDoc", "documento_desconocido")
        tipo_comprobante = nombres_comprobantes_excel.get(codDoc, "documento_desconocido")
        documento["infoTributaria"]["codDoc"] = tipo_comprobante
        sheets_data[tipo_comprobante].append(documento)

    # Creación de hojas y escritura de datos
    for sheet_name, data in sheets_data.items():
        if not data:
            continue

        if sheet_index == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        # Seleccionar el orden de las columnas apropiado para cada tipo de documento
        column_order = column_order_fac if sheet_name == "Facturas" \
                       else column_order_nc if sheet_name == "Notas de Crédito" \
                       else column_order_ret

        # Escribir encabezados
        for col_num, header in enumerate(column_order, start=1):
            ws.cell(row=1, column=col_num, value=header)

        # Escribir datos
        for row_index, doc in enumerate(data, start=2):
            # Elegir la función de aplanamiento correcta según el tipo de documento
            aplanado = aplanar_diccionario_fac(doc) if sheet_name == "Facturas" \
                       else aplanar_diccionario_nc(doc) if sheet_name == "Notas de Crédito" \
                       else aplanar_diccionario_ret(doc)

            for col_num, key in enumerate(column_order, start=1):
                valor = aplanado.get(key, "N/A")
                # Formatear datos específicos
                if key in ["totalSinImpuestos", "totalDescuento", "Base 0%", "Base 5%", "Base 12%", "Base No Objeto", "Base Exento",
                           "IVA 0%", "IVA 5%", "IVA 12%", "IVA 15%", "importeTotal", "valorModificacion", "IVA", "Renta", "ISD"] and valor != "N/A":
                    valor = float(valor)
                elif key in ["fechaEmision", "fechaEmisionDocSustento"]:
                    valor = formatear_fecha(valor)

                ws.cell(row=row_index, column=col_num, value=valor)
            id_receptor = aplanado["ruc_receptor"]

        sheet_index += 1  # Incrementa el contador después de procesar una hoja
        adjust_column_width(ws)
    fecha_creacion = datetime.now().strftime("%d-%m-%Y")
    file_name = f"{id_receptor}_reporte_contabilidad_{fecha_creacion}.xlsx"
    # Guardar el archivo
    wb.save(file_path + file_name)
    print(f"Datos guardados en el archivo {file_name}")